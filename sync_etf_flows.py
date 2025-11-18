#!/usr/bin/env python3
"""
ETF Flow Data Synchronization Script

This script automates the daily synchronization of ETF flow data from a Bloomberg
export file to multiple tracking sheets.

Installation:
    pip install -r requirements.txt

Usage:
    python sync_etf_flows.py

Author: Expert Python Developer
Date: 2025-11-17
"""

import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import sys
from typing import Dict, Any, Optional
import yfinance as yf


def create_flow_lookup(source_file: str) -> Dict[str, float]:
    """
    Read the source Bloomberg export Excel file and create a ticker-to-flow lookup map.

    This function reads ALL sheets from the source file and combines them into a single
    dictionary mapping ticker symbols to their 1D flow values.

    Args:
        source_file: Path to the source Excel file (e.g., "ETF1_uf4xn3oe.xlsx")

    Returns:
        Dictionary mapping ticker symbols to flow values (e.g., {"SPY US": -1195.342})

    Raises:
        FileNotFoundError: If the source file doesn't exist
        KeyError: If required columns are not found in the file
    """
    print(f"\n[INFO] Reading source file: {source_file}")

    if not Path(source_file).exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")

    flow_map = {}

    try:
        # Read ALL sheets from the Excel file
        all_sheets = pd.read_excel(source_file, sheet_name=None, engine='openpyxl')

        print(f"[INFO] Found {len(all_sheets)} sheet(s) in source file")

        for sheet_name, df in all_sheets.items():
            print(f"[INFO] Processing sheet: {sheet_name}")

            # Check if required columns exist
            if 'Ticker' not in df.columns:
                print(f"[WARNING] 'Ticker' column not found in sheet '{sheet_name}', skipping...")
                continue

            # Look for the flow column - it should be named " (M USD)" with a leading space
            # We'll try multiple variations to be robust
            flow_column = None
            for col in df.columns:
                if '(M USD)' in str(col) and 'Flow' in str(col):
                    flow_column = col
                    break
                elif col == ' (M USD)':
                    flow_column = col
                    break

            if flow_column is None:
                print(f"[WARNING] Flow column ' (M USD)' not found in sheet '{sheet_name}', skipping...")
                continue

            print(f"[INFO] Using flow column: '{flow_column}'")

            # Process each row
            row_count = 0
            for idx, row in df.iterrows():
                ticker = row.get('Ticker')
                flow_value = row.get(flow_column)

                # Skip rows without a ticker or with "Median" as ticker
                if pd.isna(ticker) or str(ticker).strip() == '' or str(ticker).strip() == 'Median':
                    continue

                # Skip rows without a valid flow value
                if pd.isna(flow_value):
                    continue

                # Convert flow value to float
                try:
                    flow_value = float(flow_value)
                    flow_map[str(ticker).strip()] = flow_value
                    row_count += 1
                except (ValueError, TypeError):
                    print(f"[WARNING] Invalid flow value for ticker '{ticker}': {flow_value}")
                    continue

            print(f"[INFO] Extracted {row_count} ticker(s) from sheet '{sheet_name}'")

        if not flow_map:
            print("[ERROR] No valid ticker data found in any sheet!")
            return {}

        print(f"[SUCCESS] Created flow lookup map with {len(flow_map)} total ticker(s)")
        return flow_map

    except Exception as e:
        print(f"[ERROR] Failed to read source file: {str(e)}")
        raise


def parse_multiplier_from_column_name(column_name: str) -> float:
    """
    Parse the multiplier from a column name based on leverage indicators.

    Examples:
        "UPRO (3x L)" -> 3.0 (Long 3x)
        "SPXS (3x S)" -> -3.0 (Short 3x)
        "AGQ(2x L)" -> 2.0 (Long 2x)
        "ZSL(2x S)" -> -2.0 (Short 2x)
        "TSL(1.25x L)" -> 1.25 (Long 1.25x)
        "Regular Column" -> 1.0 (No multiplier)

    Args:
        column_name: The column name to parse

    Returns:
        The multiplier (positive for Long, negative for Short)
    """
    import re

    column_lower = column_name.lower()

    # Look for pattern like "(3x L)" or "(3x S)" or "(1.25x L)"
    # Match patterns: (NUMBERx L) or (NUMBERx S)
    pattern = r'\((\d+(?:\.\d+)?)\s*x\s*([ls])\)'
    match = re.search(pattern, column_lower)

    if match:
        multiplier = float(match.group(1))
        direction = match.group(2)  # 'l' for long, 's' for short

        if direction == 's':
            return -multiplier
        else:
            return multiplier

    # No multiplier found, return 1.0
    return 1.0


def get_vwap_ticker_for_sheet(sheet_name: str) -> Optional[str]:
    """
    Get the ticker symbol to use for VWAP calculation for a given sheet.

    Args:
        sheet_name: Name of the Excel sheet

    Returns:
        Ticker symbol (e.g., "SPY", "QQQ", "TSLA") or None if not found
    """
    # Mapping of sheet names to ticker symbols for VWAP
    vwap_mapping = {
        'S&P 500 ETF': 'ES=F',
        'Nasdaq 100 ETF': 'NQ=F',
        'Russel 2000 ETF': 'RTY=F',
        'Bonds': 'TLT',
        'Gold ETF': 'GC=F',
        'Silver ETF': 'SI=F',
        'Brent ETF': 'BZ=F',
        'Natural Gas': 'NG=F',
        'Palladium ETF': 'PA=F',
        'Platinum ETF': 'PL=F',
        'Copper ETF': 'HG=F',
        'SEMIC': 'SMH',
        'NVDA': 'NVDA',
        'AVGO': 'AVGO',
        'TSLA': 'TSLA',
        'META': 'META',
        'AAPL': 'AAPL',
        'MSFT': 'MSFT',
        'GOOG': 'GOOG',
        'PANW': 'PANW',
        'IBIT': 'IBIT',
        'ETHA': 'ETHA',
        'SOL': 'SOL-USD',
        'BOFA': 'BAC',
    }

    return vwap_mapping.get(sheet_name)


def calculate_vwap(df: pd.DataFrame) -> Optional[float]:
    """
    Calculate VWAP (Volume Weighted Average Price) for a DataFrame.
    Similar to pandas_ta.vwap() but simplified.

    VWAP = Σ(Typical_Price × Volume) / Σ(Volume)
    where Typical_Price = (High + Low + Close) / 3

    Args:
        df: DataFrame with OHLCV data

    Returns:
        Last VWAP value or None if calculation fails
    """
    try:
        if df.empty or 'Volume' not in df.columns:
            return None

        # Make an explicit copy to avoid SettingWithCopyWarning
        df = df.copy()

        # Calculate Typical Price (HLC/3)
        df['Typical_Price'] = (df['High'] + df['Low'] + df['Close']) / 3

        # Calculate cumulative (Typical_Price * Volume) and cumulative Volume
        df['TP_Volume'] = df['Typical_Price'] * df['Volume']
        df['Cum_TP_Volume'] = df['TP_Volume'].cumsum()
        df['Cum_Volume'] = df['Volume'].cumsum()

        # Calculate VWAP
        df['VWAP'] = df['Cum_TP_Volume'] / df['Cum_Volume']

        # Return the last VWAP value
        return df['VWAP'].iloc[-1]

    except Exception as e:
        return None


def fetch_vwap_for_date(ticker: str, date: str) -> Optional[float]:
    """
    Fetch VWAP (Volume Weighted Average Price) for a ticker on a specific date using yfinance.
    Uses custom cumulative VWAP calculation.

    Args:
        ticker: Ticker symbol (e.g., "SPY", "TSLA", "GC=F", "SI=F")
        date: Date in YYYY-MM-DD format

    Returns:
        VWAP value or None if data not available
    """
    try:
        # Convert date string to datetime
        target_date = datetime.strptime(date, "%Y-%m-%d")
        stock = yf.Ticker(ticker)

        # Try to fetch intraday data for better VWAP calculation
        try:
            df = stock.history(start=date,
                              end=(target_date + timedelta(days=1)).strftime("%Y-%m-%d"),
                              interval='5m')

            if not df.empty and len(df) > 1:
                df = df.dropna()

                # Calculate VWAP using custom function
                vwap_value = calculate_vwap(df)

                if vwap_value is not None:
                    return round(vwap_value, 2)
        except:
            pass  # Fall back to daily data

        # Fallback: Use daily data if intraday is not available
        start_date = target_date - timedelta(days=5)
        end_date = target_date + timedelta(days=1)

        df = stock.history(start=start_date.strftime("%Y-%m-%d"),
                          end=end_date.strftime("%Y-%m-%d"))

        if df.empty:
            print(f"[WARNING] No price data found for {ticker} on {date}")
            return None

        # Find the row for our target date
        df.index = pd.to_datetime(df.index).date
        target_date_obj = target_date.date()

        if target_date_obj not in df.index:
            print(f"[WARNING] No price data for {ticker} on {date} (market closed?)")
            return None

        row = df.loc[target_date_obj]

        # For daily data, use typical price
        vwap = (row['High'] + row['Low'] + row['Close']) / 3

        return round(vwap, 2)

    except Exception as e:
        print(f"[ERROR] Failed to fetch VWAP for {ticker} on {date}: {str(e)}")
        return None


def calculate_adjusted_total_flow(df: pd.DataFrame, row_idx: int, job_config: Dict[str, Any]) -> float:
    """
    Calculate the Adjusted Total Flow for a row by summing all flow columns except VWAP.
    Note: Values in columns are already multiplied by their leverage factors, so just sum them.

    Args:
        df: DataFrame containing the sheet data
        row_idx: Index of the row to calculate for
        job_config: Job configuration containing column information

    Returns:
        The sum of all flow values (already multiplied, excluding VWAP column)
    """
    job_type = job_config['type']
    total = 0.0

    if job_type == 'complex':
        # For complex types, sum all mapped columns (already multiplied)
        mapping = job_config.get('mapping', {})
        for column_name in mapping.keys():
            if column_name in df.columns:
                value = df.at[row_idx, column_name]
                if pd.notna(value):
                    try:
                        total += float(value)
                    except (ValueError, TypeError):
                        pass
    elif job_type == 'simple':
        # For simple types, just get the single flow column value (already multiplied)
        flow_column = job_config.get('flow_column')
        if flow_column and flow_column in df.columns:
            value = df.at[row_idx, flow_column]
            if pd.notna(value):
                try:
                    total = float(value)
                except (ValueError, TypeError):
                    pass

    return total


def update_statistics_table(df: pd.DataFrame, sheet_name: str, adjusted_total_col: str, vwap_col: Optional[str]) -> pd.DataFrame:
    """
    Update the statistics table (LAST DAY, LAST 5 DAYS, LAST 20 DAYS) in the sheet.

    Args:
        df: DataFrame containing the sheet data
        sheet_name: Name of the sheet being processed
        adjusted_total_col: Name of the adjusted total flow column
        vwap_col: Name of the VWAP column (None for simple sheets without separate VWAP)

    Returns:
        Updated DataFrame with statistics
    """
    # Find rows with valid date and flow data (exclude statistics rows)
    valid_data_mask = df['Date'].notna() & pd.to_datetime(df['Date'], errors='coerce').notna()
    valid_data = df[valid_data_mask].copy()

    if len(valid_data) == 0:
        return df

    # Filter to only rows with valid adjusted total flow values
    if adjusted_total_col in valid_data.columns:
        valid_data = valid_data[valid_data[adjusted_total_col].notna()].copy()

    if len(valid_data) == 0:
        return df

    # Get the last N days of data
    last_day_data = valid_data.iloc[-1:] if len(valid_data) >= 1 else None
    last_5_days_data = valid_data.iloc[-5:] if len(valid_data) >= 5 else valid_data
    last_20_days_data = valid_data.iloc[-20:] if len(valid_data) >= 20 else valid_data

    # Calculate statistics
    stats = {}

    if last_day_data is not None and adjusted_total_col in valid_data.columns:
        stats['last_day_flow'] = last_day_data[adjusted_total_col].iloc[0]

        # For VWAP, find the most recent row with a valid VWAP value
        if vwap_col and vwap_col in df.columns:
            valid_vwap_data = valid_data[valid_data[vwap_col].notna()]
            if len(valid_vwap_data) > 0:
                stats['last_day_vwap'] = valid_vwap_data[vwap_col].iloc[-1]

    if adjusted_total_col in valid_data.columns:
        stats['last_5_days_flow'] = last_5_days_data[adjusted_total_col].sum()
        stats['last_20_days_flow'] = last_20_days_data[adjusted_total_col].sum()

        if vwap_col and vwap_col in valid_data.columns:
            # Find Product column
            product_col = None
            for col in df.columns:
                if 'product' in str(col).lower():
                    product_col = col
                    break

            # Calculate weighted average VWAP using Product column
            # Weighted Average VWAP = Sum(Products) / Sum(Adjusted Totals)
            if product_col and product_col in valid_data.columns:
                # For LAST 5 DAYS
                last_5_products = last_5_days_data[last_5_days_data[product_col].notna()][product_col]
                last_5_flows = last_5_days_data[last_5_days_data[product_col].notna()][adjusted_total_col]

                if len(last_5_products) > 0 and last_5_flows.sum() != 0:
                    stats['last_5_days_vwap'] = last_5_products.sum() / last_5_flows.sum()
                else:
                    stats['last_5_days_vwap'] = 0.0

                # For LAST 20 DAYS
                last_20_products = last_20_days_data[last_20_days_data[product_col].notna()][product_col]
                last_20_flows = last_20_days_data[last_20_days_data[product_col].notna()][adjusted_total_col]

                if len(last_20_products) > 0 and last_20_flows.sum() != 0:
                    stats['last_20_days_vwap'] = last_20_products.sum() / last_20_flows.sum()
                else:
                    stats['last_20_days_vwap'] = 0.0
            else:
                # Fallback to simple average if Product column not found
                last_5_vwap = last_5_days_data[last_5_days_data[vwap_col].notna()][vwap_col]
                last_20_vwap = last_20_days_data[last_20_days_data[vwap_col].notna()][vwap_col]

                stats['last_5_days_vwap'] = last_5_vwap.mean() if len(last_5_vwap) > 0 else 0.0
                stats['last_20_days_vwap'] = last_20_vwap.mean() if len(last_20_vwap) > 0 else 0.0

    # Find and update statistics table rows
    # Look for cells containing "LAST DAY", "LAST 5 DAYS", "LAST 20 DAYS"
    for idx, row in df.iterrows():
        for col in df.columns:
            cell_value = str(row[col]).strip() if pd.notna(row[col]) else ""

            if cell_value == "LAST DAY":
                # Find the flow and vwap columns (usually the next two columns)
                col_idx = df.columns.get_loc(col)
                if col_idx + 1 < len(df.columns) and 'last_day_flow' in stats:
                    df.at[idx, df.columns[col_idx + 1]] = stats['last_day_flow']
                if col_idx + 2 < len(df.columns) and 'last_day_vwap' in stats:
                    df.at[idx, df.columns[col_idx + 2]] = stats['last_day_vwap']

            elif cell_value == "LAST 5 DAYS":
                col_idx = df.columns.get_loc(col)
                if col_idx + 1 < len(df.columns) and 'last_5_days_flow' in stats:
                    df.at[idx, df.columns[col_idx + 1]] = stats['last_5_days_flow']
                if col_idx + 2 < len(df.columns) and 'last_5_days_vwap' in stats:
                    df.at[idx, df.columns[col_idx + 2]] = stats['last_5_days_vwap']

            elif cell_value == "LAST 20 DAYS":
                col_idx = df.columns.get_loc(col)
                if col_idx + 1 < len(df.columns) and 'last_20_days_flow' in stats:
                    df.at[idx, df.columns[col_idx + 1]] = stats['last_20_days_flow']
                if col_idx + 2 < len(df.columns) and 'last_20_days_vwap' in stats:
                    df.at[idx, df.columns[col_idx + 2]] = stats['last_20_days_vwap']

    print(f"[INFO] Updated statistics table for {sheet_name}")
    return df


def process_file(job_config: Dict[str, Any], flow_map: Dict[str, float], target_date: str) -> None:
    """
    Process a single destination file and update it with new flow data.

    This function implements the "upsert" logic:
    - If a row for the target date exists, UPDATE the flow values in place
    - If no row exists for the target date, APPEND a new row with the date and flow values

    Args:
        job_config: Configuration dictionary containing file_path, type, and mapping/ticker info
        flow_map: Dictionary mapping ticker symbols to flow values
        target_date: Target date in YYYY-MM-DD format (usually yesterday's date)

    Raises:
        FileNotFoundError: If the destination file doesn't exist
        KeyError: If required columns are not found
    """
    file_path = job_config['file_path']
    job_type = job_config['type']

    print(f"\n[INFO] Processing: {file_path}")

    # Parse the file path to extract workbook name and sheet name
    # Format: "Flows-2.xlsx - S&P 500 ETF.csv"
    # We need to extract: workbook="Flows-2.xlsx", sheet="S&P 500 ETF"

    if ' - ' in file_path:
        parts = file_path.split(' - ', 1)
        workbook_name = parts[0]
        sheet_name = parts[1].replace('.csv', '').strip()
    else:
        print(f"[ERROR] Invalid file path format: {file_path}")
        return

    if not Path(workbook_name).exists():
        raise FileNotFoundError(f"Destination workbook not found: {workbook_name}")

    try:
        # Read the specific sheet from the workbook
        df = pd.read_excel(workbook_name, sheet_name=sheet_name, engine='openpyxl')

        # Check if 'Date' column exists
        if 'Date' not in df.columns:
            print(f"[ERROR] 'Date' column not found in sheet '{sheet_name}'")
            return

        # Store original Date column format before conversion
        original_dates = df['Date'].copy()

        # Convert Date column to datetime with flexible parsing (handles DD.MM.YYYY, YYYY-MM-DD, etc.)
        # dayfirst=True handles European date format like 17.11.2025
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)

        # Convert target_date to datetime
        target_date_dt = pd.to_datetime(target_date)

        # Check if a row for the target date already exists (comparing datetime, will match regardless of time)
        # Normalize both to date only for comparison
        existing_row_mask = df['Date'].dt.date == target_date_dt.date()
        row_exists = existing_row_mask.any()

        if job_type == 'complex':
            # Complex type: Multiple columns to update based on mapping
            mapping = job_config.get('mapping', {})

            if not mapping:
                print(f"[ERROR] No mapping found for complex job: {file_path}")
                return

            updates_made = 0

            if row_exists:
                # UPDATE existing row
                print(f"[INFO] Updating existing row for date: {target_date}")
                row_idx = df[existing_row_mask].index[0]

                for column_name, ticker in mapping.items():
                    if column_name not in df.columns:
                        print(f"[WARNING] Column '{column_name}' not found in sheet, skipping...")
                        continue

                    if ticker in flow_map:
                        # Get raw value and multiply by leverage factor
                        raw_value = flow_map[ticker]
                        multiplier = parse_multiplier_from_column_name(column_name)
                        multiplied_value = raw_value * multiplier
                        df.at[row_idx, column_name] = multiplied_value
                        updates_made += 1
                    else:
                        print(f"[WARNING] Ticker '{ticker}' not found in flow map, using 0.0")
                        df.at[row_idx, column_name] = 0.0

                print(f"[INFO] Updated {updates_made} column(s)")
            else:
                # APPEND new row
                print(f"[INFO] Appending new row for date: {target_date}")

                # Create a new row with all columns set to pd.NA initially
                new_row = {col: pd.NA for col in df.columns}
                new_row['Date'] = target_date_dt

                # Fill in the mapped columns
                for column_name, ticker in mapping.items():
                    if column_name in df.columns:
                        if ticker in flow_map:
                            # Get raw value and multiply by leverage factor
                            raw_value = flow_map[ticker]
                            multiplier = parse_multiplier_from_column_name(column_name)
                            multiplied_value = raw_value * multiplier
                            new_row[column_name] = multiplied_value
                            updates_made += 1
                        else:
                            print(f"[WARNING] Ticker '{ticker}' not found in flow map, using 0.0")
                            new_row[column_name] = 0.0

                # Append the new row using loc to avoid FutureWarning
                df.loc[len(df)] = new_row
                print(f"[INFO] Appended row with {updates_made} value(s)")

        elif job_type == 'simple':
            # Simple type: Single column to update with one ticker
            ticker = job_config.get('ticker')
            flow_column = job_config.get('flow_column')

            if not ticker or not flow_column:
                print(f"[ERROR] Missing ticker or flow_column for simple job: {file_path}")
                return

            if flow_column not in df.columns:
                print(f"[ERROR] Flow column '{flow_column}' not found in sheet '{sheet_name}'")
                return

            raw_value = flow_map.get(ticker)
            if raw_value is None:
                print(f"[WARNING] Ticker '{ticker}' not found in flow map, using 0.0")
                raw_value = 0.0

            # Apply multiplier to the value before storing
            multiplier = parse_multiplier_from_column_name(flow_column)
            flow_value = raw_value * multiplier

            if row_exists:
                # UPDATE existing row
                print(f"[INFO] Updating existing row for date: {target_date}")
                row_idx = df[existing_row_mask].index[0]
                df.at[row_idx, flow_column] = flow_value
                print(f"[INFO] Updated column '{flow_column}' with value: {flow_value}")
            else:
                # APPEND new row
                print(f"[INFO] Appending new row for date: {target_date}")

                # Create a new row with all columns set to pd.NA initially
                new_row = {col: pd.NA for col in df.columns}
                new_row['Date'] = target_date_dt
                new_row[flow_column] = flow_value

                # Append the new row using loc to avoid FutureWarning
                df.loc[len(df)] = new_row
                print(f"[INFO] Appended row with value: {flow_value}")

        else:
            print(f"[ERROR] Unknown job type: {job_type}")
            return

        # Fetch and fill VWAP value for this date
        # Find VWAP column
        # First try to find column with "vwap" in name
        # If not found, try to find column matching sheet name (for NVDA, GOOG, AAPL, PANW, etc.)
        vwap_col = None
        for col in df.columns:
            # Find VWAP column but exclude Product column
            if 'vwap' in str(col).lower() and 'product' not in str(col).lower():
                vwap_col = col
                break

        # If not found, check if column name matches sheet name (for ticker-specific sheets)
        if not vwap_col:
            for col in df.columns:
                if str(col).strip().upper() == sheet_name.upper():
                    vwap_col = col
                    break

        if vwap_col:
            # Get the ticker for VWAP
            vwap_ticker = get_vwap_ticker_for_sheet(sheet_name)
            if vwap_ticker:
                vwap_value = fetch_vwap_for_date(vwap_ticker, target_date)
                if vwap_value is not None:
                    if row_exists:
                        row_idx = df[existing_row_mask].index[0]
                    else:
                        row_idx = len(df) - 1  # The newly appended row

                    df.at[row_idx, vwap_col] = vwap_value
                    print(f"[INFO] Set {vwap_col} = {vwap_value} (from {vwap_ticker})")

        # Calculate and fill Adjusted Total Flow for the current row
        # Find the Adjusted Total Flow column (could be named differently per sheet)
        adjusted_total_col = None

        for col in df.columns:
            if 'adjusted total' in str(col).lower() or (job_type == 'simple' and str(col).lower() == 'flow'):
                adjusted_total_col = col

        if adjusted_total_col:
            if row_exists:
                row_idx = df[existing_row_mask].index[0]
            else:
                row_idx = len(df) - 1  # The newly appended row

            # Calculate the adjusted total flow for this row
            adjusted_total_value = calculate_adjusted_total_flow(df, row_idx, job_config)
            df.at[row_idx, adjusted_total_col] = adjusted_total_value
            print(f"[INFO] Set {adjusted_total_col} = {adjusted_total_value}")

        # Calculate and fill Product column (VWAP × Adjusted Total)
        product_col = None
        for col in df.columns:
            if 'product' in str(col).lower():
                product_col = col
                break

        if product_col and vwap_col and adjusted_total_col:
            if row_exists:
                row_idx = df[existing_row_mask].index[0]
            else:
                row_idx = len(df) - 1  # The newly appended row

            # Get VWAP and Adjusted Total values
            vwap_value = df.at[row_idx, vwap_col]
            adjusted_total_value = df.at[row_idx, adjusted_total_col]

            # Calculate Product = VWAP × Adjusted Total
            if pd.notna(vwap_value) and pd.notna(adjusted_total_value):
                product_value = float(vwap_value) * float(adjusted_total_value)
                df.at[row_idx, product_col] = product_value
                print(f"[INFO] Set {product_col} = {product_value}")

        # Update statistics table (LAST DAY, LAST 5 DAYS, LAST 20 DAYS)
        # Re-find vwap_col for statistics update (same logic as above)
        vwap_col_for_stats = None
        for col in df.columns:
            if 'vwap' in str(col).lower() and 'product' not in str(col).lower():
                vwap_col_for_stats = col
                break

        # If not found, check if column name matches sheet name (for ticker-specific sheets)
        if not vwap_col_for_stats:
            for col in df.columns:
                if str(col).strip().upper() == sheet_name.upper():
                    vwap_col_for_stats = col
                    break

        if adjusted_total_col:
            df = update_statistics_table(df, sheet_name, adjusted_total_col, vwap_col_for_stats)

        # Convert Date column to string format "DD MMM YYYY" (e.g., "17 Nov 2025")
        # But preserve original format for certain sheets
        preserve_date_format_sheets = ['BOFA']

        if 'Date' in df.columns and sheet_name not in preserve_date_format_sheets:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%d %b %Y')

        # Save the updated DataFrame back to the Excel sheet
        # We need to read all sheets, update the specific one, and write back
        preserve_date_format_sheets = ['BOFA']

        with pd.ExcelFile(workbook_name, engine='openpyxl') as xls:
            all_sheets = {}
            for sheet in xls.sheet_names:
                # For sheets with preserved date format, read Date column as string
                if sheet in preserve_date_format_sheets:
                    all_sheets[sheet] = pd.read_excel(xls, sheet_name=sheet, dtype={'Date': str})
                else:
                    all_sheets[sheet] = pd.read_excel(xls, sheet_name=sheet)

        # Update the specific sheet
        all_sheets[sheet_name] = df

        # Write all sheets back to the workbook (dates are now strings)
        with pd.ExcelWriter(workbook_name, engine='openpyxl', mode='w') as writer:
            for sheet, data in all_sheets.items():
                data.to_excel(writer, sheet_name=sheet, index=False)

        print(f"[SUCCESS] Saved changes to: {workbook_name} (sheet: {sheet_name})")

    except Exception as e:
        print(f"[ERROR] Failed to process file: {str(e)}")
        raise


def format_statistics_table_in_sheet(workbook_path: str, sheet_name: str) -> None:
    """
    Apply formatting to the statistics table in an individual sheet to match ALL sheet style.

    Args:
        workbook_path: Path to the Excel workbook
        sheet_name: Name of the sheet to format
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        wb = load_workbook(workbook_path)

        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]

        # Define styles matching ALL sheet
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        label_font = Font(bold=True, italic=True, size=10)
        header_font = Font(italic=True, size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Find and format statistics rows
        statistics_labels = ["LAST DAY", "LAST 5 DAYS", "LAST 20 DAYS"]

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value).strip() if cell.value else ""

                # Format statistics labels
                if cell_value in statistics_labels:
                    # Format the label cell
                    cell.font = label_font
                    cell.alignment = center_align
                    cell.border = thin_border

                    # Format the flow value (next column)
                    flow_cell = ws.cell(row=row_idx, column=col_idx + 1)
                    if flow_cell.value is not None and flow_cell.value != '':
                        try:
                            flow_val = float(flow_cell.value)
                            flow_cell.alignment = right_align
                            flow_cell.border = thin_border
                            flow_cell.number_format = '#,##0.00'
                            # Apply conditional formatting
                            flow_cell.fill = green_fill if flow_val >= 0 else red_fill
                        except (ValueError, TypeError):
                            flow_cell.alignment = right_align
                            flow_cell.border = thin_border

                    # Format VWAP/Average cell (next next column)
                    vwap_cell = ws.cell(row=row_idx, column=col_idx + 2)
                    if vwap_cell.value is not None and vwap_cell.value != '':
                        try:
                            vwap_cell.alignment = right_align
                            vwap_cell.border = thin_border
                            vwap_cell.number_format = '#,##0.00'
                        except:
                            pass

        wb.save(workbook_path)
        print(f"[INFO] Applied table formatting to {sheet_name}")

    except Exception as e:
        print(f"[WARNING] Could not format statistics table in {sheet_name}: {str(e)}")


def create_all_statistics_sheet(workbook_path: str, sheet_names: list) -> None:
    """
    Create an 'ALL' sheet with statistics dashboard for all tickers in grid layout.

    Args:
        workbook_path: Path to the Excel workbook
        sheet_names: List of sheet names to collect statistics from
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"\n[INFO] Creating ALL statistics dashboard...")

    # Read all sheets to collect statistics
    all_stats = []

    # Exclude these sheets from ALL
    exclude_sheets = ['IBIT', 'SOL', 'ETHA', 'BOFA']

    with pd.ExcelFile(workbook_path, engine='openpyxl') as xls:
        for sheet_name in sheet_names:
            # Skip excluded sheets
            if sheet_name in exclude_sheets:
                continue

            try:
                df = pd.read_excel(xls, sheet_name=sheet_name)

                # Don't skip sheets without Date column - they might still have statistics
                stats_data = {
                    'ticker': sheet_name,
                    'last_day_flow': None,
                    'last_day_vwap': None,
                    'last_5_flow': None,
                    'last_5_vwap': None,
                    'last_20_flow': None,
                    'last_20_vwap': None
                }

                for idx, row in df.iterrows():
                    for col in df.columns:
                        cell_value = str(row[col]).strip() if pd.notna(row[col]) else ""
                        col_idx = df.columns.get_loc(col)

                        if cell_value == "LAST DAY":
                            if col_idx + 1 < len(df.columns):
                                stats_data['last_day_flow'] = row[df.columns[col_idx + 1]]
                            if col_idx + 2 < len(df.columns):
                                stats_data['last_day_vwap'] = row[df.columns[col_idx + 2]]

                        elif cell_value == "LAST 5 DAYS":
                            if col_idx + 1 < len(df.columns):
                                stats_data['last_5_flow'] = row[df.columns[col_idx + 1]]
                            if col_idx + 2 < len(df.columns):
                                stats_data['last_5_vwap'] = row[df.columns[col_idx + 2]]

                        elif cell_value == "LAST 20 DAYS":
                            if col_idx + 1 < len(df.columns):
                                stats_data['last_20_flow'] = row[df.columns[col_idx + 1]]
                            if col_idx + 2 < len(df.columns):
                                stats_data['last_20_vwap'] = row[df.columns[col_idx + 2]]

                all_stats.append(stats_data)

            except Exception as e:
                print(f"[WARNING] Could not read statistics from {sheet_name}: {str(e)}")
                continue

    if not all_stats:
        print("[WARNING] No statistics found to create ALL sheet")
        return

    # Create workbook and ALL sheet
    wb = load_workbook(workbook_path)

    if 'ALL' in wb.sheetnames:
        del wb['ALL']

    ws = wb.create_sheet('ALL', 0)

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ticker_font = Font(bold=True, italic=True, color="FF0000", size=11)  # Red italic
    header_font = Font(italic=True, bold=False, size=10)
    label_font = Font(bold=True, italic=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Grid layout: 4 columns of tables
    tables_per_row = 4
    table_width = 3  # columns per table
    table_height = 4  # rows per table
    col_spacing = 1  # space between tables
    row_spacing = 1  # space between table rows

    for idx, stats in enumerate(all_stats):
        # Calculate position in grid
        grid_col = idx % tables_per_row
        grid_row = idx // tables_per_row

        # Starting cell position
        start_col = 1 + grid_col * (table_width + col_spacing)
        start_row = 1 + grid_row * (table_height + row_spacing)

        # Row 1: Ticker name and headers
        # Cell A: Ticker name
        cell = ws.cell(row=start_row, column=start_col)
        cell.value = stats['ticker']
        cell.font = ticker_font
        cell.alignment = center_align
        cell.border = thin_border

        # Cell B: "FLOW" header
        cell = ws.cell(row=start_row, column=start_col + 1)
        cell.value = "FLOW"
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # Cell C: "AVERAGE" header
        cell = ws.cell(row=start_row, column=start_col + 2)
        cell.value = "AVERAGE"
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # Row 2: LAST DAY
        cell = ws.cell(row=start_row + 1, column=start_col)
        cell.value = "LAST DAY"
        cell.font = label_font
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=start_row + 1, column=start_col + 1)
        cell.value = stats['last_day_flow']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''
        if cell.value is not None:
            cell.fill = green_fill if float(cell.value) >= 0 else red_fill

        cell = ws.cell(row=start_row + 1, column=start_col + 2)
        cell.value = stats['last_day_vwap']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''

        # Row 3: LAST 5 DAYS
        cell = ws.cell(row=start_row + 2, column=start_col)
        cell.value = "LAST 5 DAYS"
        cell.font = label_font
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=start_row + 2, column=start_col + 1)
        cell.value = stats['last_5_flow']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''
        if cell.value is not None:
            cell.fill = green_fill if float(cell.value) >= 0 else red_fill

        cell = ws.cell(row=start_row + 2, column=start_col + 2)
        cell.value = stats['last_5_vwap']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''

        # Row 4: LAST 20 DAYS
        cell = ws.cell(row=start_row + 3, column=start_col)
        cell.value = "LAST 20 DAYS"
        cell.font = label_font
        cell.alignment = center_align
        cell.border = thin_border

        cell = ws.cell(row=start_row + 3, column=start_col + 1)
        cell.value = stats['last_20_flow']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''
        if cell.value is not None:
            cell.fill = green_fill if float(cell.value) >= 0 else red_fill

        cell = ws.cell(row=start_row + 3, column=start_col + 2)
        cell.value = stats['last_20_vwap']
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '#,##0.00' if cell.value else ''

    # Set column widths
    for col in range(1, tables_per_row * (table_width + col_spacing) + 1):
        if col % (table_width + col_spacing) == 1:  # Ticker column
            ws.column_dimensions[get_column_letter(col)].width = 18
        elif col % (table_width + col_spacing) == 2:  # Flow column
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col % (table_width + col_spacing) == 3:  # Average column
            ws.column_dimensions[get_column_letter(col)].width = 12

    wb.save(workbook_path)
    print(f"[SUCCESS] Created ALL statistics dashboard with {len(all_stats)} ticker(s)")


def main():
    """
    Main execution function.

    This function orchestrates the entire synchronization process:
    1. Reads the source file and creates the flow lookup map
    2. Gets today's date
    3. Processes each destination file according to its job configuration
    """
    print("="*80)
    print("ETF FLOW DATA SYNCHRONIZATION SCRIPT")
    print("="*80)

    # Configuration - Directory paths
    SOURCE_DIR = Path("Source")
    DESTINATION_DIR = Path("Destination")

    # Validate directories exist
    if not SOURCE_DIR.exists():
        print(f"[ERROR] Source directory not found: {SOURCE_DIR}")
        sys.exit(1)

    if not DESTINATION_DIR.exists():
        print(f"[ERROR] Destination directory not found: {DESTINATION_DIR}")
        sys.exit(1)

    # Find source Excel file in Source directory
    source_files = list(SOURCE_DIR.glob("*.xlsx"))
    if not source_files:
        print(f"[ERROR] No Excel files (.xlsx) found in {SOURCE_DIR}")
        sys.exit(1)

    SOURCE_FILE = str(source_files[0])
    print(f"[INFO] Using source file: {SOURCE_FILE}")

    if len(source_files) > 1:
        print(f"[WARNING] Multiple Excel files found in {SOURCE_DIR}, using first one: {SOURCE_FILE}")

    # Find destination Excel file in Destination directory
    dest_files = list(DESTINATION_DIR.glob("*.xlsx"))
    if not dest_files:
        print(f"[ERROR] No Excel files (.xlsx) found in {DESTINATION_DIR}")
        sys.exit(1)

    DESTINATION_FILE = dest_files[0].name
    print(f"[INFO] Using destination file: {DESTINATION_FILE}")

    if len(dest_files) > 1:
        print(f"[WARNING] Multiple Excel files found in {DESTINATION_DIR}, using first one: {DESTINATION_FILE}")

    # Master configuration list for all 24 jobs
    ALL_JOBS = [
        # --- Complex (Multi-Column) Files ---
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - S&P 500 ETF.csv",
            "type": "complex",
            "mapping": {
                "IVV ": "IVV US",
                "SPY ": "SPY US",
                "UPRO (3x L)": "UPRO US",
                "SPXL (3x L)": "SPXL US",
                "SPXS (3x S)": "SPXS US",
                "SPXU (3x S)": "SPXU US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Nasdaq 100 ETF.csv",
            "type": "complex",
            "mapping": {
                "QQQ": "QQQ US",
                "QQQM": "QQQM US",
                "TQQQ (3x L)": "TQQQ US",
                "SQQQ (3x S)": "SQQQ US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Russel 2000 ETF.csv",
            "type": "complex",
            "mapping": {
                "IWM": "IWM US",
                "VTWO": "VTWO US",
                "TNA (3x L)": "TNA US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Bonds.csv",
            "type": "complex",
            "mapping": {
                "TLT": "TLT US",
                "TMF (3x L)": "TMF US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Gold ETF.csv",
            "type": "complex",
            "mapping": {
                "GLD": "GLD US",
                "IAU": "IAU US",
                "GLDM": "GLDM US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Silver ETF.csv",
            "type": "complex",
            "mapping": {
                "SLV ": "SLV US",
                "SIVR": "SIVR US",
                "AGQ(2x L)": "AGQ US",
                "ZSL(2x S)": "ZSL US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Brent ETF.csv",
            "type": "complex",
            "mapping": {
                "BNO": "BNO US",
                "DBO": "DBO US",
                "USO": "USO US",
                "UCO(2x L)": "UCO US",
                "SCO(2x S)": "SCO US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Natural Gas.csv",
            "type": "complex",
            "mapping": {
                "BOIL": "BOIL US",
                "FCG": "FCG US",
                "KOLD(2x S)": "KOLD US",
                "UNG": "UNG US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Platinum ETF.csv",
            "type": "complex",
            "mapping": {
                "PPLT": "PPLT US",
                "PLTM": "PLTM US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - SEMIC.csv",
            "type": "complex",
            "mapping": {
                "SMH": "SMH US",
                "SOXX": "SOXX US",
                "SOXS (3X S)": "SOXS US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - NVDA.csv",
            "type": "complex",
            "mapping": {
                "NVDL (2x L)": "NVDL US",
                "NVD (2x S)": "NVD US",
                "NVDX(2x L)": "NVDX US",
                "NVDU(2x L)": "NVDU US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - AVGO.csv",
            "type": "complex",
            "mapping": {
                "AVGG(2x L)": "AVGG US",
                "AVGU(2x L)": "AVGU US",
                "AVGX(2x L)": "AVGX US",
                "AVL(2x L)": "AVL US",
                "AVS(1xS)": "AVDS US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - TSLA.csv",
            "type": "complex",
            "mapping": {
                "TSLQ (2x S)": "TSLQ US",
                " TSLT (2x L)": "TSLT US",
                "TSLL (2x L)": "TSLL US",
                "TSL(1.25x L)": "TSL US",
                "TSLR(2x L)": "TSLR US",
                "TSDD(2x S)": "TSDD US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - META.csv",
            "type": "complex",
            "mapping": {
                "METU (2x L)": "METU US",
                "METD (1x S)": "METD US",
                "FBL (2x L)": "FBL US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - AAPL.csv",
            "type": "complex",
            "mapping": {
                "AAPU (2x L)": "AAPU US",
                "AAPD (1x S)": "AAPD US",
                "AAPB(2x L) ": "AAPB US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - MSFT.csv",
            "type": "complex",
            "mapping": {
                "MSFL (2x L)": "MSFL US",
                "MSFU (2x L)": "MSFU US",
                "MSFX (2x L)": "MSFX US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - GOOG.csv",
            "type": "complex",
            "mapping": {
                "GGLL (2x L)": "GGLL US",
                "GGLS (1x S)": "GGLS US",
            }
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - PANW.csv",
            "type": "complex",
            "mapping": {
                "PALU (2x L)": "PALU US",
                "PANG (2X L)": "PANG US",
            }
        },

        # --- Simple (Single-Column) Files ---
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Copper ETF.csv",
            "type": "simple",
            "ticker": "CPER US",
            "flow_column": "CPER"
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - Palladium ETF.csv",
            "type": "simple",
            "ticker": "PALL US",
            "flow_column": "PALL"
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - IBIT.csv",
            "type": "simple",
            "ticker": "IBIT US",
            "flow_column": "Flow"
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - ETHA.csv",
            "type": "simple",
            "ticker": "ETHA US",
            "flow_column": "Flow"
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - SOL.csv",
            "type": "simple",
            "ticker": "SOL US",
            "flow_column": "Flow"
        },
        {
            "file_path": f"{DESTINATION_DIR / DESTINATION_FILE} - BOFA.csv",
            "type": "simple",
            "ticker": "BOFA US",
            "flow_column": "BOFA"
        },
    ]

    # Get yesterday's date
    yesterday_date = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"\n[INFO] Processing date: {yesterday_date} (yesterday)")

    # Step 1: Create the flow lookup map from the source file
    try:
        flow_map = create_flow_lookup(SOURCE_FILE)
        if not flow_map:
            print("\n[ERROR] Failed to create flow lookup map. Exiting.")
            sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] Fatal error reading source file: {str(e)}")
        sys.exit(1)

    # Step 2: Process each destination file
    print(f"\n[INFO] Starting to process {len(ALL_JOBS)} job(s)...")

    success_count = 0
    failure_count = 0

    for idx, job in enumerate(ALL_JOBS, 1):
        print(f"\n{'='*80}")
        print(f"JOB {idx}/{len(ALL_JOBS)}")
        print(f"{'='*80}")

        try:
            process_file(job, flow_map, yesterday_date)
            success_count += 1
        except Exception as e:
            print(f"[ERROR] Job failed: {str(e)}")
            failure_count += 1
            # Continue to the next job even if this one failed
            continue

    # Create ALL statistics sheet
    try:
        # Extract sheet names from job configurations
        sheet_names = []
        for job in ALL_JOBS:
            file_path = job['file_path']
            parts = file_path.split(' - ')
            if len(parts) >= 2:
                sheet_name = parts[1].replace('.csv', '').strip()
                sheet_names.append(sheet_name)

        # Get the workbook path
        workbook_path = str(DESTINATION_DIR / DESTINATION_FILE)

        # Create the ALL statistics dashboard
        create_all_statistics_sheet(workbook_path, sheet_names)

    except Exception as e:
        print(f"[WARNING] Could not create ALL statistics sheet: {str(e)}")
        # Don't fail the entire process if ALL sheet creation fails

    # Summary
    print(f"\n{'='*80}")
    print("SYNCHRONIZATION SUMMARY")
    print(f"{'='*80}")
    print(f"Total jobs: {len(ALL_JOBS)}")
    print(f"Successful: {success_count}")
    print(f"Failed: {failure_count}")
    print(f"{'='*80}\n")

    if failure_count > 0:
        sys.exit(1)
    else:
        print("[SUCCESS] All jobs completed successfully!")
        sys.exit(0)


if __name__ == "__main__":
    main()
